from bs4 import BeautifulSoup
import urllib.request
import openpyxl
import re


def get_hse(name, bt):
    spisok = []
    d = {"hse": "ВШЭ", "hse_spb": "ВШЭ СПБ", "hse_nn": "ВШЭ НН", "hse_p": "ВШЭ П"}
    vuz = d[name]
    html = open(f"vuz/hse/{name}.html", encoding="utf-8").read()
    soup = BeautifulSoup(html, "lxml")
    page = soup.findAll("tr")
    for i in range(len(page)):
        tab = page[i].find_all("td")
        urls = []
        for j in range(len(tab)):
            if j == 0:
                nup = tab[j].text
            else:
                try:
                    urls.append(tab[j].find("a").get("href"))
                except Exception:
                    print("ош")
        print(vuz, nup)
        for j in range(len(urls)):
            if "B_Os" in urls[j]:
                forma = "О"
            elif "B_SP" in urls[j]:
                forma = "С"
            elif "B_TS" in urls[j]:
                forma = "Ц"
            elif "K_OM" in urls[j]:
                forma = "К"
            elif "B_BVI" in urls[j]:
                forma = "БВИ"
            else:
                forma = "Б"

            urllib.request.urlretrieve(urls[j], "hse.xlsx")
            wookbook = openpyxl.load_workbook("hse.xlsx")
            worksheet = wookbook.active
            do = worksheet.max_row + 1
            if forma == "БВИ":
                for y in range(38, do):
                    u = []
                    for x in range(1, worksheet.max_column):
                        c = worksheet.cell(row=y, column=x).value
                        if c is not None:
                            u.append(c)
                    if len(u) > 8:
                        snils = int("".join(re.findall(r'\d+', u[1])))
                        vybor = u[3]
                        sogl = u[2]
                        spisok.append([int(snils), 311, sogl, vybor, nup, vuz, forma])

            else:
                for y in range(38, do):
                    u = []
                    for x in range(1, worksheet.max_column):
                        c = worksheet.cell(row=y, column=x).value
                        if c is not None:
                            u.append(c)
                    if len(u) > 10:
                        try:
                            snils = int("".join(re.findall(r'\d+', u[1])))
                            ball = int(u[-2])
                            vybor = u[3]
                            sogl = u[2]
                            if ball > bt or "Б" not in forma:
                                spisok.append([int(snils), ball, sogl, vybor, nup, vuz, forma])
                            else:
                                break
                        except Exception:
                            print("Ошибка")

    return spisok